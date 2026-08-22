#!/usr/bin/env python3
"""
Official Visa Sponsorship Stats — Ireland
===========================================
Downloads Ireland's Department of Enterprise, Tourism and Employment (DETE)
"Employment Permits Issued to Companies" spreadsheets — real, official,
government-published data on which employers actually received work-permit
approvals, and how many, per year. Source pages:
  https://enterprise.gov.ie/en/publications/employment-permit-statistics-2026.html
  https://enterprise.gov.ie/en/publications/employment-permit-statistics-2024.html

This is a genuinely different (and much stronger) signal than scanning job
descriptions for sponsorship language: it's the actual count of permits DETE
issued to that employer, published monthly.

WHAT IT DOES:
  1. Downloads each year's "permits issued to companies" .xlsx
  2. Prints the column headers it detected (SANITY-CHECK THESE — gov.ie can
     change the layout between years without notice; this script guesses
     which column is the employer name and which is the permit count)
  3. Fuzzy-matches employer names in the spreadsheet against your
     ireland_job_radar_HARSHIT_MASTER.csv company list (government filings often use the full
     legal entity name, e.g. "Google Ireland Limited" rather than "Google")
  4. Writes official_permit_stats.json — company -> permits per year + total

HOW OFTEN TO RUN:
  DETE updates these files roughly monthly. Re-running job_pipeline.py every
  15 minutes does NOT need to re-run this — once a month is enough:
      python visa_stats.py

USAGE:
    pip install requests openpyxl
    python visa_stats.py --companies ireland_job_radar_HARSHIT_MASTER.csv --output official_permit_stats.json
"""

import argparse
import csv
import io
import json
import re
import sys

try:
    import requests
    from openpyxl import load_workbook
except ImportError:
    print("This script needs: pip install requests openpyxl")
    sys.exit(1)

# Confirmed working source URLs as of Aug 2026 — DETE adds a new year's page
# each January. Check https://enterprise.gov.ie/en/publications/ and add the
# new year's "permits issued to companies" .xlsx link here when it appears.
YEAR_URLS = {
    "2024": "https://enterprise.gov.ie/en/publications/publication-files/permits-issued-to-companies-2024.xlsx",
    "2025": "https://assets.gov.ie/static/documents/d629d531/permits-issued-to-companies-2025.xlsx",
    "2026": "https://enterprise.gov.ie/en/publications/publication-files/employment-permits-issued-to-companies-2026.xlsx",
}

CORP_SUFFIXES = re.compile(
    r"\b(limited|ltd|plc|unlimited company|uc|inc|incorporated|group|holdings|"
    r"ireland|international|corporation|corp|company|co)\b",
    re.IGNORECASE,
)
NON_ALNUM = re.compile(r"[^a-z0-9]+")


def normalize_name(name: str) -> str:
    """Fully concatenated, no separators — safe for EXACT matching only,
    since the whole string must agree either way. NOT safe as the basis
    for substring matching (see normalize_name_spaced below)."""
    name = re.sub(r"\([^)]*\)", " ", name)  # drop parenthetical e.g. "(Broadcom)"
    name = CORP_SUFFIXES.sub(" ", name)
    name = NON_ALNUM.sub("", name.lower())
    return name.strip()


def normalize_name_spaced(name: str) -> str:
    """Same cleanup as normalize_name, but keeps a single space between
    words instead of concatenating everything together. Needed for a real
    fix: with company names fully concatenated (no separators), a short
    name like "Visa" can silently match INSIDE an unrelated word (e.g.
    "Advisable Consulting" -> "advisableconsulting" contains "visa" at
    position 2). Preserving word boundaries here lets the partial-match
    fallback require a genuine whole-word occurrence instead of a raw
    substring anywhere in the string."""
    name = re.sub(r"\([^)]*\)", " ", name)
    name = CORP_SUFFIXES.sub(" ", name)
    name = re.sub(r"[^a-z0-9]+", " ", name.lower())
    return re.sub(r"\s+", " ", name).strip()


def download_workbook(url, session):
    resp = session.get(url, timeout=30)
    resp.raise_for_status()
    return load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)


def guess_columns(header_row):
    name_col, count_col = None, None
    for idx, cell in enumerate(header_row):
        val = str(cell).strip().lower() if cell is not None else ""
        if not val:
            continue
        if name_col is None and ("employer" in val or "company" in val):
            name_col = idx
        if count_col is None and ("permit" in val or "number" in val or "total" in val or "count" in val):
            count_col = idx
    return name_col, count_col


def parse_year(url, year, session):
    """Returns {employer_name: permit_count} for one year's workbook."""
    print(f"  Downloading {year} data...")
    try:
        wb = download_workbook(url, session)
    except Exception as e:
        print(f"    FAILED to download/parse {year}: {e}")
        return {}

    sheet = wb.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return {}

    name_col, count_col = guess_columns(header)
    print(f"    Detected header: {header}")
    print(f"    -> using column {name_col} as employer name, column {count_col} as permit count")
    if name_col is None:
        print(f"    Could not confidently find an employer-name column for {year} — skipping this year.")
        return {}

    counts = {}
    for row in rows:
        if row is None or name_col >= len(row):
            continue
        employer = row[name_col]
        if not employer or not str(employer).strip():
            continue
        employer = str(employer).strip()
        if count_col is not None and count_col < len(row) and isinstance(row[count_col], (int, float)):
            n = int(row[count_col])
        else:
            # No count column found (or file is one-row-per-permit) —
            # count occurrences of this employer instead.
            n = 1
        counts[employer] = counts.get(employer, 0) + n

    return counts


def match_companies(csv_companies, year_data):
    """year_data: {year: {employer_name: count}}. Returns per-CSV-company stats."""
    normalized_index = []  # (normalize_name, normalize_name_spaced, original_employer)
    all_employers = set()
    for counts in year_data.values():
        all_employers.update(counts.keys())
    for emp in all_employers:
        normalized_index.append((normalize_name(emp), normalize_name_spaced(emp), emp))

    results = {}
    for company in csv_companies:
        target = normalize_name(company)
        target_spaced = normalize_name_spaced(company)
        if not target:
            continue
        # Real fix: the old fallback did a raw substring check on fully
        # concatenated names, meaning a short company name like "Visa"
        # could silently match INSIDE an unrelated word (e.g. "Advisable
        # Consulting" -> "advisableconsulting" contains "visa"). Now
        # requires a genuine whole-word match on the space-preserving
        # normalization instead — "Visa" still correctly matches "Visa
        # Europe Services", but can no longer match text buried inside a
        # longer unrelated word. Threshold kept at 4 (unchanged) since the
        # word-boundary requirement is what actually fixes the collision,
        # not the length — raising the threshold further would just block
        # legitimate short names like "Visa" and "Meta" from matching at all.
        matches = [orig for norm, norm_spaced, orig in normalized_index
                   if norm == target or
                   (len(target_spaced) >= 4 and
                    re.search(rf"\b{re.escape(target_spaced)}\b", norm_spaced))]
        if not matches:
            continue

        by_year = {}
        total = 0
        for year, counts in year_data.items():
            year_total = sum(counts.get(m, 0) for m in matches)
            if year_total:
                by_year[year] = year_total
                total += year_total

        if total:
            results[company] = {
                "matched_employer_names": sorted(set(matches)),
                "permits_by_year": by_year,
                "total_permits": total,
            }
    return results


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--companies", default="ireland_job_radar_HARSHIT_MASTER.csv")
    ap.add_argument("--output", default="official_permit_stats.json")
    args = ap.parse_args()

    with open(args.companies, newline="", encoding="utf-8-sig") as f:
        csv_companies = [row["company_name"].strip() for row in csv.DictReader(f)]

    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (compatible; job-search-dashboard/1.0)"})

    print("Fetching official DETE employment permit statistics...")
    year_data = {}
    for year, url in YEAR_URLS.items():
        year_data[year] = parse_year(url, year, session)

    print("\nMatching employers to your company list...")
    matched = match_companies(csv_companies, year_data)

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(matched, f, indent=2)

    print(f"\nMatched {len(matched)} of {len(csv_companies)} companies to official permit records.")
    print(f"Saved to {args.output}")
    print("\nSANITY CHECK: skim the 'matched_employer_names' in that file — fuzzy name")
    print("matching can mismatch on very short/common company names. Remove any bad")
    print("matches by hand if you spot them; job_pipeline.py will use whatever's there.")


if __name__ == "__main__":
    main()
